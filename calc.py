import pkuseg
import os
from collections import Counter
from pprint import pprint
import pandas as pd
from openpyxl import load_workbook
seg = pkuseg.pkuseg()

MOST_USE_NUM = 20

def to_dict(self):
    return {
        'work': self.x,
        'nums': self.y,
    }

def walkFolder(file):
    for root, dirs, files in os.walk(file):
        # root 表示当前正在访问的文件夹路径
        # dirs 表示该文件夹下的子目录名list
        # files 表示该文件夹下的文件list
        # 遍历所有的文件夹
        for d in dirs:
            dir_path = os.path.join(root, d)
            if (dir_path.find('data') > 0):
                # if (dir_path == './data\Beijing' or dir_path == './data\Dalian'):
                    walkFile(dir_path)

def walkFile(folder):
    for root, dirs, files in os.walk(folder):
        for f in files:
            file_path = os.path.join(root, f)
            if (file_path.find('content')>=0):
                calcContent(folder,file_path)
            if (file_path.find('title')>=0):
                calcTitle(folder,file_path)

def calcContent(folder_path, input_file_path):
    str = open(input_file_path, 'r', encoding="utf-8").read()
    text = seg.cut(str)  # 进行分词

    stopwords = []
    with open('./hit_stopwords.txt',encoding='utf-8') as f:
        stopwords = f.read()
    new_text = []
    for w in text:
        if w not in stopwords:
            new_text.append(w)
    
    counter = Counter(new_text)
    pprint(counter.most_common(MOST_USE_NUM))
    df = pd.DataFrame(counter.most_common(MOST_USE_NUM), columns=['word','nums'])

    book = load_workbook("./counter_content.xlsx")
    writer = pd.ExcelWriter("./counter_content.xlsx", engine = 'openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=folder_path.split('\\')[1])
    writer.save()
    writer.close()
        

def calcTitle(folder_path, input_file_path):
    str = open(input_file_path, 'r', encoding="utf-8").read()
    text = seg.cut(str)  # 进行分词

    stopwords = []
    with open('./hit_stopwords.txt',encoding='utf-8') as f:
        stopwords = f.read()
    new_text = []
    for w in text:
        if w not in stopwords:
            new_text.append(w)
    
    counter = Counter(new_text)
    pprint(counter.most_common(MOST_USE_NUM))
    df = pd.DataFrame(counter.most_common(MOST_USE_NUM), columns=['word','nums'])
    
    book = load_workbook("./counter_title.xlsx")
    writer = pd.ExcelWriter("./counter_title.xlsx", engine = 'openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=folder_path.split('\\')[1])
    writer.save()
    writer.close()

if __name__ == '__main__':
    walkFolder(r"./data")